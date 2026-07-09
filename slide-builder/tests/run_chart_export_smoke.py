#!/usr/bin/env python3
"""Smoke test for export_chart_data.py (C1 — think-cell Excel export).

A brief with three slides — one with a markdown chart table, one with CSV chart
data, one with no chart — exports to .xlsx and asserts: a sheet per CHART slide
(not the no-chart one), values parsed into cells, numbers stored as numbers, and
a TBD/placeholder chart-data block is skipped.

Run:  py -3 slide-builder/tests/run_chart_export_smoke.py  (python3 on macOS/Linux)
Prints "SMOKE PASSED." on success; raises AssertionError otherwise.
"""
from __future__ import annotations

import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

HERE = Path(__file__).resolve().parent
SCRIPTS = HERE.parent / "scripts"

BRIEF = """---
client_template: X
---

## Deck-level design notes

Chart export smoke.

## Slide 1 — Revenue by quarter
**Chart type:** bar
**Chart data:**
| Quarter | Revenue |
| Q1 | 12 |
| Q2 | 18 |
| Q3 | 25 |

## Slide 2 — Text only
**Slide type:** Synthesis
**Governing thought:** No chart here.

## Slide 3 — Split (CSV)
**Chart type:** donut
**Chart data:**
Segment, Share
Core, 68
Non-core, 32

## Slide 4 — Placeholder chart
**Chart type:** line
**Chart data:**
TBD — placeholder
"""


def main() -> int:
    tmp = Path(tempfile.mkdtemp(prefix="chartexp_smoke_"))
    try:
        brief = tmp / "narrative-brief-x.md"
        brief.write_text(BRIEF, encoding="utf-8")
        out = tmp / "chart-data.xlsx"

        print("[1] export")
        r = subprocess.run([sys.executable, str(SCRIPTS / "export_chart_data.py"),
                            str(brief), "--out", str(out)], capture_output=True, text=True)
        assert r.returncode == 0, r.stderr + r.stdout
        assert out.exists(), "xlsx not written"

        print("[2] verify sheets + values")
        from openpyxl import load_workbook
        wb = load_workbook(str(out))
        names = wb.sheetnames
        # slides 1 and 3 are charts; 2 (no chart) and 4 (TBD) are excluded
        assert names == ["slide_01", "slide_03"], names
        ws1 = wb["slide_01"]
        # data starts at row 3: header row then Q1/12 ...
        assert ws1["A3"].value == "Quarter" and ws1["B3"].value == "Revenue", (ws1["A3"].value, ws1["B3"].value)
        assert ws1["A4"].value == "Q1" and ws1["B4"].value == 12, (ws1["A4"].value, ws1["B4"].value)
        assert isinstance(ws1["B4"].value, (int, float)), "numbers should be stored as numbers"
        assert ws1["B6"].value == 25, ws1["B6"].value
        ws3 = wb["slide_03"]
        assert ws3["A4"].value == "Core" and ws3["B4"].value == 68, (ws3["A4"].value, ws3["B4"].value)
        print("    ok: 2 chart sheets (slide_01 bar, slide_03 donut); values parsed; numbers numeric; text/TBD slides skipped")

        print("[3] a deck with no charts exports nothing")
        nb = tmp / "no-charts.md"
        nb.write_text("## Slide 1 — Plain\n**Chart type:** none\n**Governing thought:** x.\n", encoding="utf-8")
        r = subprocess.run([sys.executable, str(SCRIPTS / "export_chart_data.py"), str(nb)],
                           capture_output=True, text=True)
        assert r.returncode == 0 and "Nothing to export" in r.stdout, r.stdout
        print("    ok: no-chart deck reports nothing to export")

        print("\nSMOKE PASSED.")
        return 0
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


if __name__ == "__main__":
    sys.exit(main())
