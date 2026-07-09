#!/usr/bin/env python3
"""export_chart_data.py — emit an Excel datasheet of a deck's chart data so
charts can be edited in think-cell (or Excel / native PowerPoint charts).

Slide Lab DRAWS charts as shapes (add_rect + add_text), so a chart in the built
deck isn't data-editable in place. But the numbers exist in the brief's per-slide
`**Chart data:**` blocks. This reads those and writes one workbook with a sheet
per chart slide, so a consultant can open it, copy the range, and paste it into a
think-cell datasheet (or a native chart's data grid).

Input: either a built deck's out dir (reads `_meta.json` → the brief it records)
or a narrative-brief `.md` directly.

  py -3 export_chart_data.py <out_dir | brief.md> [--out chart-data.xlsx]
  (python3 on macOS/Linux)

Only slides with `**Chart type:**` != none AND a non-empty, non-TBD
`**Chart data:**` block are exported. Prints the path, or says there's nothing
to export.
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

from openpyxl import Workbook  # noqa: E402

_SLIDE_HDR = re.compile(r"^##\s+Slide\s+(\d+)\s*[—\-:]\s*(.*)$", re.MULTILINE)
_CHART_TYPE = re.compile(r"\*\*Chart type:\*\*\s*(.+)", re.IGNORECASE)
_CHART_DATA = re.compile(
    r"\*\*Chart data:\*\*\s*\n?([\s\S]+?)(?=\n\*\*[A-Z][^\n]*:\*\*|\n##\s+Slide|\Z)",
    re.IGNORECASE,
)


def _resolve_brief(arg: str) -> Path | None:
    p = Path(arg)
    if p.is_dir():
        meta = p / "_meta.json"
        if meta.exists():
            try:
                b = json.loads(meta.read_text(encoding="utf-8")).get("brief")
                if b and Path(b).exists():
                    return Path(b)
            except Exception:
                pass
        # fall back to a lone brief file in the dir
        cands = sorted(p.glob("*brief*.md")) + sorted(p.glob("adopted_brief.md"))
        return cands[0] if cands else None
    if p.suffix.lower() == ".md" and p.exists():
        return p
    return None


def _parse_rows(block: str) -> list[list[str]]:
    """Turn a chart-data text block into rows of cells. Supports markdown
    tables (| a | b |), CSV, and TSV; falls back to one cell per line."""
    rows: list[list[str]] = []
    for line in block.splitlines():
        s = line.strip()
        if not s:
            continue
        # skip markdown separator rows like |---|---|
        if set(s) <= set("|-: "):
            continue
        if "|" in s:
            cells = [c.strip() for c in s.strip("|").split("|")]
        elif "\t" in s:
            cells = [c.strip() for c in s.split("\t")]
        elif "," in s:
            cells = [c.strip() for c in s.split(",")]
        else:
            cells = [s]
        rows.append(cells)
    return rows


def _chart_slides(brief_text: str) -> list[dict]:
    """Return [{n, title, chart_type, rows}] for slides with real chart data."""
    out = []
    matches = list(_SLIDE_HDR.finditer(brief_text))
    for i, m in enumerate(matches):
        n = int(m.group(1))
        title = m.group(2).strip()
        start = m.end()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(brief_text)
        block = brief_text[start:end]
        ct_m = _CHART_TYPE.search(block)
        ct = (ct_m.group(1).strip() if ct_m else "").strip().lower()
        if not ct or ct in ("none", "n/a", "-"):
            continue
        cd_m = _CHART_DATA.search(block)
        cd = (cd_m.group(1).strip() if cd_m else "")
        if not cd or cd.upper().startswith("TBD") or "placeholder" in cd.lower():
            continue
        rows = _parse_rows(cd)
        if rows:
            out.append({"n": n, "title": title, "chart_type": ct, "rows": rows})
    return out


def main(argv) -> int:
    ap = argparse.ArgumentParser(description="Export a deck's chart data to Excel for think-cell.")
    ap.add_argument("source", help="A built deck's out dir (with _meta.json) OR a brief .md.")
    ap.add_argument("--out", default=None, help="Output .xlsx path (default: <source>/chart-data.xlsx or next to the brief).")
    args = ap.parse_args(argv)

    brief = _resolve_brief(args.source)
    if not brief:
        print(f"[error] couldn't find a brief from: {args.source} "
              f"(pass a build out dir with _meta.json, or a narrative-brief .md).",
              file=sys.stderr)
        return 2
    slides = _chart_slides(brief.read_text(encoding="utf-8"))
    if not slides:
        print("No chart slides found (no slide has a chart type + real chart data). "
              "Nothing to export.")
        return 0

    if args.out:
        out = Path(args.out)
    else:
        src = Path(args.source)
        out = (src / "chart-data.xlsx") if src.is_dir() else brief.with_name("chart-data.xlsx")

    wb = Workbook()
    wb.remove(wb.active)
    for s in slides:
        ws = wb.create_sheet(title=f"slide_{s['n']:02d}"[:31])
        # A tiny header so the consultant knows which slide/chart this is; the
        # data table starts at row 3 so the range they copy into think-cell is clean.
        ws["A1"] = f"Slide {s['n']}: {s['title']}"
        ws["A2"] = f"chart type: {s['chart_type']}"
        for r_i, row in enumerate(s["rows"], start=3):
            for c_i, val in enumerate(row, start=1):
                # numbers as numbers where possible, so think-cell/Excel treat them right
                try:
                    ws.cell(row=r_i, column=c_i, value=float(val) if val.replace(".", "", 1).replace("-", "", 1).isdigit() else val)
                except Exception:
                    ws.cell(row=r_i, column=c_i, value=val)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    print(f"[ok] wrote {out}")
    print(f"     {len(slides)} chart slide(s): " + ", ".join(f"slide {s['n']} ({s['chart_type']})" for s in slides))
    print("     Open it, copy a sheet's table (from row 3), and paste into a think-cell")
    print("     datasheet — or into a native PowerPoint chart's data grid.")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
